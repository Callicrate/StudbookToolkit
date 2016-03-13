from openpyxl import Workbook
from openpyxl import load_workbook


class excelWriter:
    """These objects can be used to write Excel formatted files. Please note
    that all data is stored in RAM until close: is called which makes this
    potentially 'difficult' to use with very large data sets.
    """

    def __init__(self, name):
        """Initialize an excelWriter: object.

        Args:
           name (str):  name of the file to create when writing (include .xlsx
           in the variable if you want correct Excel naming!)

        Returns:
           excelWriter:
        """

        self.wb = Workbook()
        self.ws = self.wb.active
        self.name = name

    def writeStudbook(self, studbook):
        """write the entire contents from a studbook: into this object.

        Args:
           studbook (studbook):  studbook: to be written to this object

        Returns:
            nothing
        """

        self._writeMultipleRows(studbook.header)

        for record in studbook.directory:
            self._writeMultipleRows(record.returnExcelFormat())

    def _writeMultipleRows(self, data):
        """append multiple rows into this object from a python list

        Args:
           data (list):  each element of data: must contain a list which represents a row of data

        Returns:
           nothing
        """

        for row in data:
            self._writeRow(row)

    # data must be a simple list
    def _writeRow(self, data):
        """append a single row into this object

        Args:
           data (list):  each element of data: represents a single column of data to be written

        Returns:
            nothing
        """

        self.ws.append(data)

    def close(self):
        """save all the data from this object into an Excel file. This may be
        poorly named, but we don't really have a close-out function yet other
        than this.

        Returns:
            nothing
        """

        self.wb.save(self.name)


class excelReader:
    """These objects can be used to read Excel formatted files.
    """

    def __init__(self, name):
        """Initialize an excelReader: object.

        Args:
           name (str):  name of the file to read from

        Returns:
           excelWriter:
        """
        self.wb = load_workbook(filename = name)
        self.sheet_ranges = self.wb[self.wb.get_sheet_names()[0]]

    def getRecordsAsList(self):
        """read all records from this file into a (2-dimensional) list. Each
        element of the list is list representing an entire row of data. In
        each internal list each element is a single column.

        FOR NOW THIS ONLY READS THE 'B' AND 'C' COLUMNS

        Returns:
           returnMe (list):
        """

        returnme = []
        i = 2
        while i <= len(self.sheet_ranges.rows):
            sire = self.sheet_ranges['B'+str(i)].value
            dam = self.sheet_ranges['C'+str(i)].value
            sd = [sire, dam]
            returnme.append(sd)
            i += 1

        return returnme